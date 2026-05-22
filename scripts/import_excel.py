#!/usr/bin/env python3
"""Eenmalige import-script: leest 'Hindernislijst en materiaallijst.xlsx'
en schrijft data/data.json.

Voer uit met:  python scripts/import_excel.py [pad/naar.xlsx]
"""
from __future__ import annotations

import json
import re
import sys
import unicodedata
import xml.etree.ElementTree as ET
import zipfile
from collections import OrderedDict
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = ROOT / "Hindernislijst en materiaallijst.xlsx"
OUT_PATH = ROOT / "data" / "data.json"

# Naam-varianten tussen 'materiaallijst' en 'hindernislijst'. Sleutel =
# naam in materiaallijst (genormaliseerd via normalize_name), waarde =
# de naam zoals die in hindernislijst staat. Pas dit gerust aan als er
# ambiguiteit opgelost moet worden.
NAME_ALIASES: dict[str, str] = {
    "brug af - brug op": "Brug op-Brug af",
    "bruf af - brug op": "Brug op-Brug af",
    "boomslinger": "Touwslinger",
    "slootnet": "Sloot-net",
    "finish boog": "Stormbaan",
}

# Sheet-naam (foto-tab) -> hindernisnaam in de hindernislijst.
# Alleen nodig als de sheet-naam afwijkt van de hindernisnaam (genormaliseerd).
PHOTO_SHEET_ALIASES: dict[str, str] = {
    "bernds special 1": "Bernd's special 1",
    "bernds special 2": "Bernd's special 2",
}


# ---------- helpers ----------

def norm_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_truthy_x(value: Any) -> bool:
    s = norm_str(value).lower()
    return s in {"x", "ja", "j", "yes", "y", "true", "1"}


def parse_qty(value: Any) -> tuple[float | str | None, str]:
    """Geeft (aantal, eenheid) terug. Aantal is float waar mogelijk,
    anders de originele string (b.v. '70m'); eenheid blijft leeg
    behalve als de cel iets zoals '70m' bevat."""
    if value is None:
        return None, ""
    if isinstance(value, (int, float)):
        return float(value), ""
    s = str(value).strip()
    if not s:
        return None, ""
    m = re.match(r"^\s*([0-9]+(?:[.,][0-9]+)?)\s*([a-zA-Z]+)?\s*$", s)
    if m:
        num = float(m.group(1).replace(",", "."))
        unit = (m.group(2) or "").lower()
        return num, unit
    return s, ""


def slug_first_name(name: str) -> str:
    """Geeft genormaliseerde voornaam terug voor matching."""
    if not name:
        return ""
    s = unicodedata.normalize("NFKD", name).encode("ascii", "ignore").decode("ascii")
    s = s.strip().split()
    if not s:
        return ""
    return s[0].lower()


# ---------- bouwteams ----------

def read_bouwteams(wb) -> list[dict]:
    ws = wb["Bouwteams"]
    teams: list[dict] = []
    for row in ws.iter_rows(values_only=True):
        cells = [norm_str(c) for c in row]
        if not any(cells):
            continue
        if cells[0].lower() == "bouwhoofd":
            continue
        bouwhoofd = cells[0]
        leden = [c for c in cells[1:] if c]
        if not bouwhoofd:
            continue
        teams.append({"bouwhoofd": bouwhoofd, "leden": leden})
    return teams


def build_team_index(teams: list[dict]) -> dict[str, str]:
    """Mapping: lookup-string (volledige naam of voornaam, lowercase) ->
    volledige naam van bouwhoofd. Bouwhoofden hebben voorrang boven leden.
    """
    idx: dict[str, str] = {}
    # eerst de bouwhoofden (krijgen voorrang)
    for t in teams:
        full = t["bouwhoofd"]
        idx[full.lower()] = full
        first = slug_first_name(full)
        if first and first not in idx:
            idx[first] = full
    # daarna leden -> bouwhoofd, alleen toevoegen als de key nog vrij is
    for t in teams:
        for lid in t["leden"]:
            full_lid = lid.strip()
            if not full_lid:
                continue
            low = full_lid.lower()
            if low not in idx:
                idx[low] = t["bouwhoofd"]
            first = slug_first_name(full_lid)
            if first and first not in idx:
                idx[first] = t["bouwhoofd"]
    return idx


def match_team(raw: str, team_idx: dict[str, str]) -> tuple[str, bool]:
    """Geeft (naam, matched) terug."""
    if not raw:
        return "", False
    raw_clean = raw.strip()
    low = raw_clean.lower()
    if low in team_idx:
        return team_idx[low], True
    first = slug_first_name(raw_clean)
    if first and first in team_idx:
        return team_idx[first], True
    # gedeeltelijke match (b.v. 'Erwin Mol' -> 'Bernd Veldhuis'? nee, gebruikt voornaam)
    for key, full in team_idx.items():
        if key in low or low in key:
            return full, True
    return raw_clean, False


# ---------- hindernissen ----------

def read_hindernissen(wb) -> list[dict]:
    ws = wb["hindernislijst"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return []

    header = rows[1]
    h6 = norm_str(header[6]).lower() if len(header) > 6 else ""
    # Nieuwe lay-out (2026): kolom 6 = "Wie bericht", 7 = "BOUWER", geen "Vervangen door".
    # Oude lay-out: 6 = "Vervangen door", 7 = "Wie bericht", 8 = "BOUWER".
    if h6 and "wie bericht" in h6 and "vervang" not in h6:
        col = {
            "nr": 0,
            "naam": 1,
            "kort": 2,
            "lang": 3,
            "v1": 4,
            "v2": 5,
            "vervangen_door": None,
            "wie_bericht": 6,
            "bouwer": 7,
            "toelichting": 8,
            "uitlegger": 9,
        }
    else:
        col = {
            "nr": 0,
            "naam": 1,
            "kort": 2,
            "lang": 3,
            "v1": 4,
            "v2": 5,
            "vervangen_door": 6,
            "wie_bericht": 7,
            "bouwer": 8,
            "toelichting": 9,
            "uitlegger": 10,
        }

    def cell(r: tuple, key: str) -> str:
        j = col[key]
        if j is None:
            return ""
        if j >= len(r):
            return ""
        return norm_str(r[j])

    obstacles: list[dict] = []
    for r in rows[2:]:
        nr = r[0] if len(r) > 0 else None
        if nr is None:
            break
        try:
            nr_int = int(nr)
        except (TypeError, ValueError):
            continue
        naam = cell(r, "naam")
        if not naam:
            continue
        obstacles.append({
            "id": nr_int,
            "nr": nr_int,
            "naam": naam,
            "kort": is_truthy_x(r[col["kort"]] if len(r) > col["kort"] else None),
            "lang": is_truthy_x(r[col["lang"]] if len(r) > col["lang"] else None),
            "vrijwilliger1": cell(r, "v1"),
            "vrijwilliger2": cell(r, "v2"),
            "vervangen_door": cell(r, "vervangen_door"),
            "wie_bericht": cell(r, "wie_bericht"),
            "bouwer_raw": cell(r, "bouwer"),
            "toelichting": cell(r, "toelichting"),
            "uitlegger": cell(r, "uitlegger"),
            "extra": norm_str(r[col["uitlegger"] + 1]) if len(r) > col["uitlegger"] + 1 else "",
        })
    return obstacles


# ---------- materialen ----------

def normalize_name(s: str) -> str:
    s = unicodedata.normalize("NFKD", s or "").encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", s).strip().lower()


def read_materialen_rows(wb) -> list[dict]:
    """Leest alle materiaalregels uit het materiaallijst-blad.
    Geeft een lijst dicts met 'nr_raw', 'hindernis_naam' (uit kolom B), de
    bouwer (kolom F) en de materiaalgegevens. Kolom F bevat in de praktijk
    de naam van het bouwhoofd en wordt als primaire bron voor 'bouwteam'
    gebruikt."""
    ws = wb["materiaallijst"]
    out: list[dict] = []
    started = False
    for row in ws.iter_rows(values_only=True):
        if not started:
            cell0 = norm_str(row[0]).lower()
            if cell0 == "nr.":
                started = True
            continue
        cell0 = norm_str(row[0])
        if cell0.lower().startswith("materiaal overig"):
            break
        if cell0.lower().startswith("bouwteams"):
            break

        hindernis_naam = norm_str(row[1])
        materiaal = norm_str(row[2])
        if not hindernis_naam:
            continue
        # Bouwer-cel altijd onthouden voor team-matching, ook als er
        # geen of slechts een 'niets'-materiaal in de regel staat.
        bouwer_cel = norm_str(row[5]) if len(row) > 5 else ""

        nr_int: int | None = None
        try:
            nr_int = int(row[0]) if row[0] is not None else None
        except (TypeError, ValueError):
            nr_int = None

        materiaal_obj: dict | None = None
        if materiaal and materiaal.lower() != "niets":
            aantal, eenheid = parse_qty(row[3])
            leverancier = norm_str(row[4])
            opmerking = norm_str(row[6]) if len(row) > 6 else ""
            materiaal_obj = {
                "naam": materiaal,
                "aantal": aantal,
                "eenheid": eenheid,
                "leverancier": leverancier,
                "opmerking": opmerking,
            }

        out.append({
            "nr": nr_int,
            "hindernis_naam": hindernis_naam,
            "bouwer_mat": bouwer_cel,
            "materiaal": materiaal_obj,
        })
    return out


_NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "rel": "http://schemas.openxmlformats.org/package/2006/relationships",
}


def extract_sheet_images(xlsx_path: Path) -> dict[str, list[tuple[str, bytes]]]:
    """Geeft per sheet-naam een lijst (filename, bytes) van ingebedde
    afbeeldingen terug. openpyxl laadt deze niet betrouwbaar, dus we lezen
    de zip-structuur direct.
    """
    out: dict[str, list[tuple[str, bytes]]] = {}
    with zipfile.ZipFile(xlsx_path) as z:
        names = set(z.namelist())

        def read_xml(p: str):
            with z.open(p) as f:
                return ET.parse(f).getroot()

        wb = read_xml("xl/workbook.xml")
        wb_rels = read_xml("xl/_rels/workbook.xml.rels")
        rid2target = {
            r.get("Id"): r.get("Target")
            for r in wb_rels.findall("rel:Relationship", _NS)
        }

        def resolve(base_xml: str, target: str) -> str:
            """Loste een relatief Target-pad op tegen het XML-bestand waarin
            de relatie staat. Werkt met ../ en /."""
            if target.startswith("/"):
                return target.lstrip("/")
            base_dir = "/".join(base_xml.split("/")[:-1])
            parts = base_dir.split("/") if base_dir else []
            for piece in target.split("/"):
                if piece == "..":
                    parts.pop()
                elif piece == ".":
                    continue
                else:
                    parts.append(piece)
            return "/".join(parts)

        for s in wb.findall("main:sheets/main:sheet", _NS):
            name = s.get("name") or ""
            rid = s.get(
                "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
            )
            if not rid or rid not in rid2target:
                continue
            sheet_target = rid2target[rid]
            sheet_xml = resolve("xl/workbook.xml", sheet_target)
            sheet_rels = sheet_xml.rsplit("/", 1)[0] + "/_rels/" + sheet_xml.rsplit("/", 1)[1] + ".rels"
            if sheet_rels not in names:
                continue
            r_root = read_xml(sheet_rels)
            drawing_targets = [
                r.get("Target")
                for r in r_root.findall("rel:Relationship", _NS)
                if "drawing" in (r.get("Type") or "")
            ]
            images: list[tuple[str, bytes]] = []
            for d in drawing_targets:
                d_path = resolve(sheet_xml, d)
                d_rels = d_path.rsplit("/", 1)[0] + "/_rels/" + d_path.rsplit("/", 1)[1] + ".rels"
                if d_rels not in names:
                    continue
                dr_root = read_xml(d_rels)
                for r in dr_root.findall("rel:Relationship", _NS):
                    if "image" in (r.get("Type") or ""):
                        img_path = resolve(d_path, r.get("Target"))
                        if img_path in names:
                            with z.open(img_path) as f:
                                images.append((img_path.split("/")[-1], f.read()))
            if images:
                out[name] = images
    return out


def write_sheet_photos(
    sheet_images: dict[str, list[tuple[str, bytes]]],
    obstacles: list[dict],
    images_root: Path,
) -> tuple[dict[int, list[str]], list[str]]:
    """Schrijft per sheet-naam de bijbehorende afbeeldingen naar
    images/obstacles/<id>/xlsx-N.<ext>. Bestaande xlsx-* bestanden worden
    eerst opgeruimd, andere foto's blijven staan. Geeft per obstacle.id
    de toegevoegde paden en een lijst sheet-namen zonder match terug.
    """
    name_to_id: dict[str, int] = {}
    for o in obstacles:
        name_to_id[normalize_name(o["naam"])] = o["id"]

    added: dict[int, list[str]] = {}
    unmatched: list[str] = []
    for sheet_name, images in sheet_images.items():
        key = normalize_name(sheet_name)
        if key in (normalize_name(k) for k in PHOTO_SHEET_ALIASES):
            key = normalize_name(PHOTO_SHEET_ALIASES[key])
        target_id = name_to_id.get(key)
        if target_id is None:
            unmatched.append(sheet_name)
            continue
        obs_dir = images_root / str(target_id)
        obs_dir.mkdir(parents=True, exist_ok=True)
        # ruim oude xlsx-* bestanden op zodat hernoemen/verwijderen schoon gaat
        for f in obs_dir.glob("xlsx-*"):
            try:
                f.unlink()
            except OSError:
                pass
        for i, (orig_name, data) in enumerate(images, start=1):
            ext = Path(orig_name).suffix.lower() or ".png"
            out_path = obs_dir / f"xlsx-{i}{ext}"
            out_path.write_bytes(data)
            added.setdefault(target_id, []).append(
                f"images/obstacles/{target_id}/{out_path.name}"
            )
    return added, unmatched


def resolve_material_row_to_obstacle(
    row: dict,
    obstacles: list[dict],
    name_buckets: dict[str, list[dict]],
    nr_index: dict[int, dict],
) -> dict | None:
    """Bepaal welke hindernis bij een materiaallijst-rij hoort. Geeft het
    obstacle terug of None."""
    key = normalize_name(row["hindernis_naam"])
    if key in NAME_ALIASES:
        key = normalize_name(NAME_ALIASES[key])
    bucket = name_buckets.get(key)
    if not bucket:
        prefix_bucket = [
            o for o in obstacles
            if normalize_name(o["naam"]).startswith(key + " ")
            or normalize_name(o["naam"]) == key
        ]
        if prefix_bucket:
            bucket = prefix_bucket
        elif row["nr"] is not None and row["nr"] in nr_index:
            bucket = [nr_index[row["nr"]]]
        else:
            return None
    if len(bucket) == 1 or row["nr"] is None:
        return bucket[0]
    return min(bucket, key=lambda o: abs(o["nr"] - row["nr"]))


def assign_materials_to_obstacles(
    obstacles: list[dict],
    mat_rows: list[dict],
) -> tuple[dict[int, list[dict]], dict[int, str], list[dict]]:
    """Koppelt materiaal-rijen aan hindernissen via naam-match (canonical id =
    obstacle.id uit de hindernislijst). Bij dubbele namen in de hindernislijst
    wordt gedisambigueerd op basis van het dichtstbijzijnde nummer in de
    materiaallijst. Niet-gematchte rijen worden teruggegeven als losse
    'unmatched' lijst.

    Geeft daarnaast per obstacle.id de eerste niet-lege 'Bouwer'-waarde
    uit de materiaallijst terug (deze is doorgaans betrouwbaarder dan de
    BOUWER-kolom uit de hindernislijst).
    """
    name_buckets: dict[str, list[dict]] = {}
    for o in obstacles:
        key = normalize_name(o["naam"])
        if not key:
            continue
        name_buckets.setdefault(key, []).append(o)
    nr_index: dict[int, dict] = {o["nr"]: o for o in obstacles}

    per_obstacle: dict[int, list[dict]] = {}
    bouwer_by_id: dict[int, str] = {}
    unmatched: list[dict] = []
    for row in mat_rows:
        target = resolve_material_row_to_obstacle(row, obstacles, name_buckets, nr_index)
        if target is None:
            if row["materiaal"] is not None:
                unmatched.append(row)
            continue
        tid = target["id"]
        if row["materiaal"] is not None:
            per_obstacle.setdefault(tid, []).append(row["materiaal"])
        if row["bouwer_mat"] and tid not in bouwer_by_id:
            bouwer_by_id[tid] = row["bouwer_mat"]
    return per_obstacle, bouwer_by_id, unmatched


# ---------- main ----------

def main() -> int:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx_path.exists():
        print(f"Bestand niet gevonden: {xlsx_path}", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    teams = read_bouwteams(wb)
    team_idx = build_team_index(teams)

    obstacles = read_hindernissen(wb)
    mat_rows = read_materialen_rows(wb)
    materialen_per_obs, bouwer_by_id, unmatched_materials = assign_materials_to_obstacles(
        obstacles, mat_rows
    )

    # Foto's uit het Excel-bestand halen en wegschrijven onder images/obstacles/<id>/.
    images_root = ROOT / "images" / "obstacles"
    sheet_images = extract_sheet_images(xlsx_path)
    _written, unmatched_photo_sheets = write_sheet_photos(sheet_images, obstacles, images_root)

    unmatched: list[tuple[int, str]] = []
    for obs in obstacles:
        raw_hl = obs.pop("bouwer_raw")
        raw_ml = bouwer_by_id.get(obs["id"], "")
        # Materiaallijst is leidend voor het bouwteam (bevat doorgaans de
        # volledige bouwhoofd-naam); hindernislijst.BOUWER is fallback.
        team, ok = ("", False)
        for candidate in (raw_ml, raw_hl):
            if candidate:
                team, ok = match_team(candidate, team_idx)
                if ok:
                    break
        obs["bouwteam"] = team
        if (raw_ml or raw_hl) and not ok:
            unmatched.append((obs["nr"], raw_ml or raw_hl))
        obs["materialen"] = materialen_per_obs.get(obs["nr"], [])
        # foto's: alles dat onder images/obstacles/<id>/ staat
        photo_dir = images_root / str(obs["id"])
        photos: list[str] = []
        if photo_dir.exists():
            for f in sorted(photo_dir.iterdir()):
                if f.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp"}:
                    photos.append(f"images/obstacles/{obs['id']}/{f.name}")
        obs["fotos"] = photos

    data = {
        "obstacles": obstacles,
        "bouwteams": teams,
        "_meta": {
            "source": xlsx_path.name,
            "unmatched_bouwers": [
                {"nr": nr, "raw": raw} for nr, raw in unmatched
            ],
            "unmatched_material_rows": [
                {
                    "hindernis_naam": r["hindernis_naam"],
                    "nr_in_materiaallijst": r["nr"],
                    "materiaal": r["materiaal"]["naam"],
                }
                for r in unmatched_materials
            ],
            "unmatched_photo_sheets": unmatched_photo_sheets,
        },
    }

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(
        json.dumps(data, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    print(f"Geschreven: {OUT_PATH} ({len(obstacles)} hindernissen, {len(teams)} bouwteams)")
    if unmatched:
        print("LET OP - niet-gematchte BOUWER waarden (handmatig bewerken in data.json):")
        for nr, raw in unmatched:
            print(f"  nr {nr}: {raw!r}")
    if unmatched_materials:
        print(
            "LET OP - materiaalregels zonder bijbehorende hindernis "
            "(niet meegenomen, alleen-obstakels modus):"
        )
        for r in unmatched_materials:
            print(
                f"  hindernis={r['hindernis_naam']!r} "
                f"(nr in materiaallijst={r['nr']}) materiaal={r['materiaal']['naam']!r}"
            )
    if unmatched_photo_sheets:
        print(
            "LET OP - foto-werkbladen zonder bijbehorende hindernis "
            "(voeg eventueel een entry toe aan PHOTO_SHEET_ALIASES):"
        )
        for s in unmatched_photo_sheets:
            print(f"  sheet={s!r}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
